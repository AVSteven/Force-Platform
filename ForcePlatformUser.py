import tkinter as tk
from tkinter import ttk
from tkinter import StringVar
from tkinter import IntVar
from datetime import date
from tkinter import messagebox
from dateutil.relativedelta import *
import matplotlib
import matplotlib.pyplot as plt
matplotlib.use("TkAgg")
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
import matplotlib.animation as animation
from matplotlib import style
import math
import random
from tinydb import TinyDB, Query
import mplcursors
from openpyxl import Workbook
from openpyxl import load_workbook
import os, sys
import serial
import time
from scipy import interpolate
from scipy.signal import butter, lfilter, freqz
import numpy as np


#########################################################################
#                                                                       #
#               Declaración de variables y constantes                   #
#                                                                       #    
#   - Se crean al iniciar el programa                                   #
#########################################################################

####################### Declaracion de Json(TinyDB) #######################
db_Clinico = TinyDB('db_Clinico.json')
dbTbSujetoClinico = db_Clinico.table('TB_SujetoClinico')
dbTbEvaluadoresClinico = db_Clinico.table('TB_Evaluadores')
dbTbSujEvalClinico = db_Clinico.table('TB_SujEvals')

db_Investigacion = TinyDB('db_Investigacion.json')
dbTbSujetoInvestigacion = db_Investigacion.table('TB_SujetoInvestigacion')
dbTbEvaluadoresInvestigacion = db_Investigacion.table('TB_Evaluadores')
dbTbSujEvalInvestigacion = db_Investigacion.table('TB_SujEvals')

#################### Fuentes texto del programa ###########################
LARGE_FONT= ("Verdana", 12)
NORM_FONT = ("Helvetica", 10)

######################### Variables de control ############################

FechaHoy = date.today()         #Fecha del sistema

flagGrabacion = False           #Indica al programa si el boton de inciar o deterne grabacion, fue presionado(True cuando esta grabando)
SeleccionSujeto = False         #Indica si se utilizao las ventans de la pestaña de grabacion, ya sea nuevo o existente(True si el usuario ingreso a esta ventana).
flagTipoInvestigacion = True    #Indica si la investigacion es clinica(True) o de investigacion(False)
flagTipoExportar = True         #Indica si se debe importar con pseudonimo o con nombre(True con nombre - False con pseudonimo)
TestTime = 0                    #Unicamente se utiliza cuando se activa el protoco de prueba
DatosFuerza = []                #Almacena los datos de fuerza cuando se inicia la grabacion
DatosTiempo = []                #Almacena los datos de tiempo cuando se inicia la grabacion

Embebido = False                #Indica si el embebido esta conectado

EstadoCalibracion = False       #Se activa en True cuando se realiza un cambio en la pestaña de calibracion

ceroDato1 = 0                   #Indican el valos de ajuste a cero, tomado de la lectura del embebido. Cuando se presiona establecer 0
ceroDato2 = 0
ceroDato3 = 0
ceroDato4 = 0

######################## Conexion del embebido ############################

try:
    ser = serial.Serial('COM3',115200, timeout = 5)
    Embebido = True
except:
    None

####################### Figuras de los graficos ###########################

x_len = 200        #Numero de puntos en X para mostrar
y_range = [0, 200]  #Rango de los valores para mostrar en Y

#Figuras utilizadsa para mostrar los graficos
fig = plt.figure(figsize=(5,2.65))
fig2 = plt.figure(figsize=(5,2.65))
figTotal = plt.figure()
ax = fig.add_subplot(1, 1, 1)
ax2 = fig2.add_subplot(1, 1, 1)
axTotal = figTotal.add_subplot(1, 1, 1)
xs = list(range(0, 200))
ys = [0] * x_len
ax.set_ylim(y_range)

#Linea que sera actualizada en la animacion
line, = ax.plot(xs, ys)

#Informacion del grafico
ax.set_xlabel('Tiempo')
ax.set_ylabel('Fuerza')

#########################################################################
#                                                                       #
#                       Funciones de prueba                             #
#                                                                       #
#   -Se pueden utilizar para verificar secciones del programa           #
#########################################################################
def imprimirlista():    #Esta función no se utiliza, a menos que se requieran verificar los datos que se ingresan a las listas correspondientes.
    print("Fuerza" + str(DatosFuerza))
    print(len(DatosFuerza))
    DatosTiempo = list(range(0, len(DatosFuerza)))
    TMP_DatosTiempo = []
    for i in DatosTiempo:
        TMP_DatosTiempo.append(round(i * 0.05,2))
    print("Tiempo" + str(TMP_DatosTiempo))
    print(len(TMP_DatosTiempo))

def randomNum():    #Cuando se activa el protocolo de prueba(PSoC no conectado) se utiliza esta funcion para verificar el funcionamiento del programa.
    global TestTime
    global EstadoCalibracion
    global DatoCalibracion_A
    global DatoCalibracion_B
    global DatoCalibracion_C
    global DatoCalibracion_D
    if(EstadoCalibracion):
        TestTime = TestTime+1
        if(TestTime<=100):
            Dato1=random.randrange(0,10)-DatoCalibracion_A
            Dato2=random.randrange(0,10)-DatoCalibracion_B
            Dato3=random.randrange(0,10)-DatoCalibracion_C
            Dato4=random.randrange(0,10)-DatoCalibracion_D
        elif(TestTime>100 and TestTime<=150):
            Dato1=random.randrange(100,125)-DatoCalibracion_A
            Dato2=random.randrange(100,125)-DatoCalibracion_B
            Dato3=random.randrange(100,125)-DatoCalibracion_C
            Dato4=random.randrange(100,125)-DatoCalibracion_D
        elif(TestTime>150 and TestTime<=200):
            Dato1=random.randrange(0,25)-DatoCalibracion_A
            Dato2=random.randrange(0,25)-DatoCalibracion_B
            Dato3=random.randrange(0,25)-DatoCalibracion_C
            Dato4=random.randrange(0,25)-DatoCalibracion_D
        elif(TestTime>200 and TestTime<=250):
            Dato1=0
            Dato2=0
            Dato3=0
            Dato4=0
        elif(TestTime>250 and TestTime<=300):
            Dato1=random.randrange(500,625)-DatoCalibracion_A
            Dato2=random.randrange(500,625)-DatoCalibracion_B
            Dato3=random.randrange(500,625)-DatoCalibracion_C
            Dato4=random.randrange(500,625)-DatoCalibracion_D
        elif(TestTime>300 and TestTime<=350):
            Dato1=random.randrange(250,350)-DatoCalibracion_A
            Dato2=random.randrange(250,350)-DatoCalibracion_B
            Dato3=random.randrange(250,350)-DatoCalibracion_C
            Dato4=random.randrange(250,350)-DatoCalibracion_D
        else:
            Dato1=random.randrange(100,125)-DatoCalibracion_A
            Dato2=random.randrange(100,125)-DatoCalibracion_B
            Dato3=random.randrange(100,125)-DatoCalibracion_C
            Dato4=random.randrange(100,125)-DatoCalibracion_D
            
        DatoFuerza=Dato1+Dato2+Dato3+Dato4
        DatoFiltrado = 0
        if(DatoFuerza<=CalibracionMagnitud):
            DatoFiltrado=DatoFuerza
            DatoFuerza = 0

        strDatoSensor1.set(Dato1)
        strDatoSensor2.set(Dato2)
        strDatoSensor3.set(Dato3)
        strDatoSensor4.set(Dato4)
        strDatoFuerza.set(DatoFuerza)
        strDatoFiltrado.set(DatoFiltrado)
        return DatoFuerza
    else:
        TestTime = TestTime+1
        if(TestTime<=100):
            Dato1=random.randrange(0,10)
            Dato2=random.randrange(0,10)
            Dato3=random.randrange(0,10)
            Dato4=random.randrange(0,10)
        elif(TestTime>100 and TestTime<=150):
            Dato1=random.randrange(100,125)
            Dato2=random.randrange(100,125)
            Dato3=random.randrange(100,125)
            Dato4=random.randrange(100,125)
        elif(TestTime>150 and TestTime<=200):
            Dato1=random.randrange(0,25)
            Dato2=random.randrange(0,25)
            Dato3=random.randrange(0,25)
            Dato4=random.randrange(0,25)
        elif(TestTime>200 and TestTime<=250):
            Dato1=0
            Dato2=0
            Dato3=0
            Dato4=0
        elif(TestTime>250 and TestTime<=300):
            Dato1=random.randrange(500,625)
            Dato2=random.randrange(500,625)
            Dato3=random.randrange(500,625)
            Dato4=random.randrange(500,625)
        elif(TestTime>300 and TestTime<=350):
            Dato1=random.randrange(250,350)
            Dato2=random.randrange(250,350)
            Dato3=random.randrange(250,350)
            Dato4=random.randrange(250,350)
        else:
            Dato1=random.randrange(100,125)
            Dato2=random.randrange(100,125)
            Dato3=random.randrange(100,125)
            Dato4=random.randrange(100,125)

        DatoFuerza=Dato1+Dato2+Dato3+Dato4
        DatoFiltrado = 0
        if(DatoFuerza<=5):
            DatoFiltrado=DatoFuerza
            DatoFuerza = 0
        

        strDatoSensor1.set(Dato1)
        strDatoSensor2.set(Dato2)
        strDatoSensor3.set(Dato3)
        strDatoSensor4.set(Dato4)
        strDatoFuerza.set(DatoFuerza)
        strDatoFiltrado.set(DatoFiltrado)
        return DatoFuerza



#########################################################################
#                                                                       #
#               Funciones ventanas de informacion                       #
#                                                                       #
#   -   Carga el archivo txt que contiene la información que se         #
#       muetra al usuario.                                              #
#   -   Carga las imagenes utilizadas para ilustrar la información      #
#                                                                       #
#########################################################################
def InformacionPlataformaFuerza():
    vInformacionPlataformaFuerza = tk.Toplevel()
    frame_P1 = tk.Frame(vInformacionPlataformaFuerza)
    frame_P1.pack(fill="x",side=tk.TOP)
    frame_P2 = tk.Frame(vInformacionPlataformaFuerza)
    frame_P2.pack(fill="x",side=tk.TOP)

    lblTitulo = tk.Label(frame_P1, text="----- Plataforma de fuerza biomecánica -----", font=LARGE_FONT)
    lblTitulo.pack(side=tk.TOP,pady=10,padx=10)

    fileInfo=open("Informacion/Informacion.txt", "r")
    InfoPlataforma = ""
    if fileInfo.mode == 'r':
        contenido =fileInfo.read()
        InfoPlataforma = contenido.split("$")[0]
    lblTxtPlataforma = tk.Label(frame_P2, text=InfoPlataforma, font=NORM_FONT,padx = 10,justify=tk.LEFT,wraplength=500)
    lblTxtPlataforma.pack(side=tk.LEFT,pady=10,padx=10)

    ImgPlataforma = tk.PhotoImage(file="Imagenes/Plataforma.PNG")
    lblImgPlataforma = tk.Label(frame_P2, image=ImgPlataforma)
    lblImgPlataforma.photo = ImgPlataforma
    lblImgPlataforma.pack(side=tk.RIGHT,pady=10,padx=10)

def InformacionErgoTec():
    vInformacionPlataformaFuerza = tk.Toplevel()
    frame_P1 = tk.Frame(vInformacionPlataformaFuerza)
    frame_P1.pack(fill="x",side=tk.TOP)
    frame_P2 = tk.Frame(vInformacionPlataformaFuerza)
    frame_P2.pack(fill="x",side=tk.TOP)

    lblTitulo = tk.Label(frame_P1, text="----- Laboratorio de ergonomía aplicada(ErgoTec) -----", font=LARGE_FONT)
    lblTitulo.pack(side=tk.TOP,pady=10,padx=10)

    fileInfo=open("Informacion/Informacion.txt", "r")
    InfoPlataforma = ""
    if fileInfo.mode == 'r':
        contenido =fileInfo.read()
        InfoPlataforma = contenido.split("$")[1]
    lblTxtPlataforma = tk.Label(frame_P2, text=InfoPlataforma, font=NORM_FONT,padx = 10,justify=tk.LEFT,wraplength=500)
    lblTxtPlataforma.pack(side=tk.LEFT,pady=10,padx=10)

    ImgPlataforma = tk.PhotoImage(file="Imagenes/Laboratorio.PNG")
    lblImgPlataforma = tk.Label(frame_P2, image=ImgPlataforma)
    lblImgPlataforma.photo = ImgPlataforma
    lblImgPlataforma.pack(side=tk.RIGHT,pady=10,padx=10)
    

#########################################################################
#                                                                       #
#               Funciones para el control del programa                  #
#                                                                       #
#   -La mayoría se llaman por medio de los botones de la interfaz       #
#########################################################################

def LecturaSensores():
###########################################
##           Dato de entrada             ##
## b'A 1.528 B 1.539 C 1.532 D 1.520 \n' ##  
###########################################
    if(EstadoCalibracion):
        line = ser.readline()
        try:
            #Dato1=float(line[2:7])-DatoCalibracion_A-ceroDato1+0.7
            Dato1=(54.49640*float(line[2:7])-109.42877)-DatoCalibracion_A-ceroDato1
##            Dato1=float(line[2:7])
        except:
            Dato1=0
        try:
            #Dato2=float(line[9:15])-DatoCalibracion_B-ceroDato2+0.7
            Dato2=(48.48*float(line[9:15])-96.96)-DatoCalibracion_B-ceroDato2
##            Dato2=float(line[9:15])
        except:
            Daro2=0
        try:
            #Dato3=float(line[18:23])-DatoCalibracion_C-ceroDato3+0.7
            Dato3=(53.48631*float(line[18:23])-107.88190)-DatoCalibracion_C-ceroDato3
##            Dato3=float(line[18:23])
        except:
            Dato3=0
        try:
            #Dato4=float(line[26:33])-DatoCalibracion_D-ceroDato4+0.7
            Dato4=(71.21034*float(line[26:33])-143.20399)-DatoCalibracion_D-ceroDato4
##            Dato4=float(line[26:33])
        except:
            Dato4=0
        #PromVoltaje=(Dato1+Dato2+Dato3+Dato4)/4
        #DatoFuerza=41667*PromVoltaje-113708
        DatoFuerza=(Dato1+Dato2+Dato3+Dato4)/4
        DatoFiltrado = 0
        if(CalibracionMagnitud == 0): 
            if(DatoFuerza<=5):
                DatoFiltrado = DatoFuerza
                DatoFuerza = 0
        else:
            if(DatoFuerza<=CalibracionMagnitud):
                DatoFiltrado = DatoFuerza
                DatoFuerza = 0
        

        strDatoSensor1.set(round(Dato1,3))
        strDatoSensor2.set(round(Dato2,3))
        strDatoSensor3.set(round(Dato3,3))
        strDatoSensor4.set(round(Dato4,3))
        strDatoFuerza.set(round(DatoFuerza,3))
        strDatoFiltrado.set(round(DatoFiltrado,3))
        return DatoFuerza
    else:
        line = ser.readline()
        try:
##            Dato1=float(line[2:7])-ceroDato1+0.7
            Dato1=(54.49640*float(line[2:7])-109.42877)-ceroDato1
##            Dato1=float(line[2:7])
        except:
            Dato1=0
        try:
##            Dato2=float(line[9:15])-ceroDato2+0.7
            Dato2=(48.48*float(line[9:15])-96.96)-ceroDato2
##            Dato2=float(line[9:15])
        except:
            Daro2=0
        try:
##            Dato3=float(line[18:23])-ceroDato3+0.7
            Dato3=(53.48631*float(line[18:23])-107.88190)-ceroDato3
##            Dato3=float(line[18:23])
        except:
            Dato3=0
        try:
##            Dato4=float(line[26:33])-ceroDato4+0.7
            Dato4=(71.21034*float(line[26:33])-143.20399)-ceroDato4
##            Dato4=float(line[26:33])
        except:
            Dato4=0
##        PromVoltaje=(Dato1+Dato2+Dato3+Dato4)/4
##        DatoFuerza=41667*PromVoltaje-113708
        DatoFuerza=(Dato1+Dato2+Dato3+Dato4)/4
        DatoFiltrado = 0
        if(DatoFuerza<=5):
            DatoFiltrado = DatoFuerza
            DatoFuerza = 0

        strDatoSensor1.set(round(Dato1,3))
        strDatoSensor2.set(round(Dato2,3))
        strDatoSensor3.set(round(Dato3,3))
        strDatoSensor4.set(round(Dato4,3))
        strDatoFuerza.set(round(DatoFuerza,3))
        strDatoFiltrado.set(DatoFiltrado)
        return DatoFuerza

def FlagGrabacion():
    global flagGrabacion
    if (flagGrabacion==True):
        strBtnGrabacion.set("Iniciar Grabación")
        flagGrabacion = False
    else:
        strBtnGrabacion.set("Detener Grabación")
        ani.event_source.start()
        flagGrabacion = True
    
def ExportarDatos():
    global SeleccionSujeto
    global flagTipoInvestigacion
    global flagTipoExportar
    global DatosFuerza
    global DatosTiempo
    if(SeleccionSujeto == True):
        wb=Workbook()
        if(flagTipoExportar==True):
            tmpFilePath = strDatoNombre.get()+"_"+strDatoApellido.get()+".xlsx"
        else:
            tmpFilePath = strDatoSujeto.get()+".xlsx"
        filePathExcel = tmpFilePath
        wb.save(filePathExcel)
        DatosTiempo = list(range(0, len(DatosFuerza)))
        TMP_DatosTiempo = []
        for i in DatosTiempo:
            TMP_DatosTiempo.append(round(i * 0.05,2))
        if(EstadoCalibracion):
            try:
                for i in DatosTiempo:
                    TMP_DatosTiempo.append(round(i * (1/CalibracionFrecuencia),2))
            except:
                for i in DatosTiempo:
                    TMP_DatosTiempo.append(round(i * 0.05,2))
        else:    
            for i in DatosTiempo:
                TMP_DatosTiempo.append(round(i * 0.05,2))        
        filePathExcel=tmpFilePath
        wb=load_workbook(filePathExcel)
        sheet=wb.active
        for row in zip(TMP_DatosTiempo, DatosFuerza):
            sheet.append(row) 
        wb.save(filePathExcel)
        DatosFuerza = []
        DatosTiempo = []
    else:
        wb=Workbook()
        filePathExcel="general.xlsx"
        wb.save(filePathExcel)
        DatosTiempo = list(range(0, len(DatosFuerza)))
        TMP_DatosTiempo = []
        for i in DatosTiempo:
            TMP_DatosTiempo.append(round(i * 0.05,2))
        if(EstadoCalibracion):
            try:
                for i in DatosTiempo:
                    TMP_DatosTiempo.append(round(i * (1/CalibracionFrecuencia),2))
            except:
                for i in DatosTiempo:
                    TMP_DatosTiempo.append(round(i * 0.05,2))
        else:    
            for i in DatosTiempo:
                TMP_DatosTiempo.append(round(i * 0.05,2))
        filePathExcel="general.xlsx"
        wb=load_workbook(filePathExcel)
        sheet=wb.active
        for row in zip(TMP_DatosTiempo, DatosFuerza):
            sheet.append(row) 
        wb.save(filePathExcel)
        DatosFuerza = []
        DatosTiempo = []

def AnimarGrafico(i, ys):
    # Leer datos del USB
    if(Embebido):
        FuerzaPlataforma = round(LecturaSensores(), 3)
    else:
        FuerzaPlataforma = round(randomNum(), 3)
    # Agregamos y a la lista
    ys.append(FuerzaPlataforma)
    if (flagGrabacion==True): 
        DatosFuerza.append(FuerzaPlataforma)
        
    ys = ys[-x_len:]

    # Actualizar la linea con los nuevos valores de Y
    line.set_ydata(ys)
    if(EstadoCalibracion):
        try:
            ani.event_source.interval = (1/CalibracionFrecuencia)*1000
        except:
            ani.event_source.interval = 50
    else:
        ani.event_source.interval = 50
    return line,

########################################################################
def StopGrafico():
    ani.event_source.stop()

def PlayGrafico():
    ani.event_source.start()

########################################################################
def MostrarGrafico():
    vMostrarGrafico = tk.Tk()
    global DatosFuerza
    global DatosTiempo
    strDatoGrafico = StringVar(vMostrarGrafico)
    frame_P1 = tk.Frame(vMostrarGrafico)
    frame_P1.pack(fill="x",side=tk.TOP)
    frame_P2 = tk.Frame(vMostrarGrafico)
    frame_P2.pack(fill="x",side=tk.TOP)
    lblInfoGrafico = tk.Label(frame_P1, text="Comentario: ", font=LARGE_FONT)
    lblInfoGrafico.pack(side=tk.LEFT,pady=10,padx=10)
    entInfoGrafico = ttk.Entry(frame_P1,textvariable=strDatoGrafico, width=150)
    entInfoGrafico.pack(side=tk.LEFT,pady=10,padx=10)
    
    DatosTiempo = list(range(0, len(DatosFuerza)))
    TMP_DatosTiempo = []
    if(EstadoCalibracion):
        try:
            for i in DatosTiempo:
                TMP_DatosTiempo.append(round(i * (1/CalibracionFrecuencia),2))
        except:
            for i in DatosTiempo:
                TMP_DatosTiempo.append(round(i * 0.05,2))
    else:    
        for i in DatosTiempo:
            TMP_DatosTiempo.append(round(i * 0.05,2))

    f = interpolate.interp1d(TMP_DatosTiempo, DatosFuerza, kind="linear")
    x_int = np.linspace(TMP_DatosTiempo[0],TMP_DatosTiempo[-1], 20)
    y_int = f(x_int)

    def FiltroPasoBajasAux(F_Corte, fs, Orden):
        tmpFs = 0.5 * fs
        FrecuanciaNormalizada = F_Corte / tmpFs
        b, a = butter(Orden, FrecuanciaNormalizada, btype='low', analog=False)
        return b, a

    def FiltroPasoBajas(Datos, F_Corte, fs, Orden):
        b, a = FiltroPasoBajasAux(F_Corte, fs, Orden)
        y = lfilter(b, a, Datos)
        return y

    if(EstadoCalibracion):
        try:
            y = FiltroPasoBajas(DatosFuerza, CalibracionFrecCorte, 50, CalibracionOrden)
            print("si")
        except:
            y = FiltroPasoBajas(DatosFuerza, 5, 50, 2)
            print("no")
    else:
        y = FiltroPasoBajas(DatosFuerza, 5, 50, 2)
        
    
    canvas = FigureCanvasTkAgg(figTotal, vMostrarGrafico)
    canvas.draw()
    canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.X, expand=True)
        
    toolbar = NavigationToolbar2Tk(canvas, vMostrarGrafico)
    toolbar.update()
    canvas._tkcanvas.pack(side=tk.TOP, fill=tk.X, expand=True)

    axTotal.clear()
    axTotal.plot(TMP_DatosTiempo, DatosFuerza)
    axTotal.plot(x_int, y_int, color="red")
    axTotal.plot(TMP_DatosTiempo, y, color="green")
    axTotal.set_title("Gráfico de fuerza vs tiempo de la plataforma de fuerza")
    axTotal.set_xlabel('Tiempo(s)')
    axTotal.set_ylabel('Fuerza(N)')
    mplcursors.cursor(multiple=True).connect("add", lambda sel: sel.annotation.set_text(entInfoGrafico.get()))
########################################################################
def EstablecerCero():
    global ceroDato1
    global ceroDato2
    global ceroDato3
    global ceroDato4
    print(ceroDato1)
    print(ceroDato2)
    print(ceroDato3)
    print(ceroDato4)
    if(float(strDatoSensor1.get())<=ceroDato1):
        ceroDato1 = ceroDato1
    else:
        ceroDato1 = float(strDatoSensor1.get())
    if(float(strDatoSensor2.get())<=ceroDato2):
        ceroDato2 = ceroDato2
    else:
        ceroDato2 = float(strDatoSensor2.get())
    if(float(strDatoSensor3.get())<=ceroDato3):
        ceroDato3 = ceroDato3
    else:
        ceroDato3 = float(strDatoSensor3.get())
    if(float(strDatoSensor4.get())<=ceroDato4):
        ceroDato4 = ceroDato4
    else:
        ceroDato4 = float(strDatoSensor4.get())
########################################################################

def Calibracion():
    vCalibracion = tk.Tk()
    
    strSensor_A = tk.DoubleVar(vCalibracion)
    strSensor_B = tk.DoubleVar(vCalibracion)
    strSensor_C = tk.DoubleVar(vCalibracion)
    strSensor_D = tk.DoubleVar(vCalibracion)
    strFrecuencia = tk.DoubleVar(vCalibracion)
    strMagnitud = tk.DoubleVar(vCalibracion)
    strFrecCorte = tk.DoubleVar(vCalibracion)
    strOrden = tk.DoubleVar(vCalibracion)

    def salirCalibracion():
        vCalibracion.destroy()

    def setDatosCalibracion():
        global EstadoCalibracion
        global DatoCalibracion_A
        global DatoCalibracion_B
        global DatoCalibracion_C
        global DatoCalibracion_D
        global CalibracionFrecuencia
        global CalibracionMagnitud
        global CalibracionFrecCorte
        global CalibracionOrden
        DatoCalibracion_A = float(strSensor_A.get())
        DatoCalibracion_B = float(strSensor_B.get())
        DatoCalibracion_C = float(strSensor_C.get())
        DatoCalibracion_D = float(strSensor_D.get())

        CalibracionFrecuencia = float(strFrecuencia.get())
        CalibracionMagnitud = float(strMagnitud.get())
        CalibracionFrecCorte = float(strFrecCorte.get())
        CalibracionOrden = float(strOrden.get())
        if(DatoCalibracion_A == 0 and DatoCalibracion_B == 0 and DatoCalibracion_C == 0 and DatoCalibracion_D == 0 and CalibracionFrecuencia == 0
           and CalibracionMagnitud == 0 and CalibracionFrecCorte == 0 and CalibracionOrden == 0):
            EstadoCalibracion = False
        else:
            EstadoCalibracion = True
        vCalibracion.destroy()
        
    
    lblInfoCalibracion = ttk.Label(vCalibracion, text="-------------------------------- Matriz de calibración --------------------------------",font=NORM_FONT)
    lblInfoCalibracion.grid(row=0, column=0,columnspan=4)
    lblSensor_A = ttk.Label(vCalibracion, text="Sensor A")
    lblSensor_A.grid(row=1, column=0)
    entSensor_A = ttk.Entry(vCalibracion,textvariable=strSensor_A)
    entSensor_A.grid(row=1, column=1,sticky='we')
    lblSensor_B = ttk.Label(vCalibracion, text="Sensor B")
    lblSensor_B.grid(row=1, column=2)
    entSensor_B = ttk.Entry(vCalibracion,textvariable=strSensor_B)
    entSensor_B.grid(row=1, column=3,sticky='we')
    lblSensor_C = ttk.Label(vCalibracion, text="Sensor C")
    lblSensor_C.grid(row=2, column=0)
    entSensor_C = ttk.Entry(vCalibracion,textvariable=strSensor_C)
    entSensor_C.grid(row=2, column=1,sticky='we')
    lblSensor_D = ttk.Label(vCalibracion, text="Sensor D")
    lblSensor_D.grid(row=2, column=2)
    entSensor_D = ttk.Entry(vCalibracion,textvariable=strSensor_D)
    entSensor_D.grid(row=2, column=3,sticky='we')
    lblInfoFrecuencia = ttk.Label(vCalibracion, text="------------- Frecuencia de muestreo / nivel de magnitud -------------",font=NORM_FONT)
    lblInfoFrecuencia.grid(row=3, column=0,columnspan=4)
    lblFrecuencia = ttk.Label(vCalibracion, text="Frecuencia")
    lblFrecuencia.grid(row=4, column=0)
    entFrecuencia = ttk.Entry(vCalibracion,textvariable=strFrecuencia)
    entFrecuencia.grid(row=4, column=1,sticky='we')
    lblMagnitud = ttk.Label(vCalibracion, text="Lectura mínima")
    lblMagnitud.grid(row=4, column=2)
    entMagnitud = ttk.Entry(vCalibracion,textvariable=strMagnitud)
    entMagnitud.grid(row=4, column=3,sticky='we')
    lblInfoFiltro = ttk.Label(vCalibracion, text="-------------------------------- Filtro paso bajas --------------------------------",font=NORM_FONT)
    lblInfoFiltro.grid(row=5, column=0,columnspan=4)
    lblFrecCorte = ttk.Label(vCalibracion, text="Frecuencia de corte")
    lblFrecCorte.grid(row=6, column=0)
    entFrecCorte = ttk.Entry(vCalibracion,textvariable=strFrecCorte)
    entFrecCorte.grid(row=6, column=1,sticky='we')
    lblOrden = ttk.Label(vCalibracion, text="Orden")
    lblOrden.grid(row=6, column=2)
    entOrden = ttk.Entry(vCalibracion,textvariable=strOrden)
    entOrden.grid(row=6, column=3,sticky='we')

    btnAceptar = ttk.Button(vCalibracion, text = "Aceptar", command = setDatosCalibracion)
    btnAceptar.grid(row=7,sticky='we',column=1)
    btnSalir = ttk.Button(vCalibracion, text = "Cancelar", command = salirCalibracion)
    btnSalir.grid(row=7,sticky='we',column=2)
    
########################################################################

def NuevoSujeto(NewExistSeleccion):
    vNuevoSujeto = tk.Tk()
    global SeleccionSujeto
    SeleccionSujeto = True
    global strDatoSujeto
    strDatoSujeto = StringVar(vNuevoSujeto)
    global flagTipoExportar
    flagTipoExportar=True

    strNombre = StringVar(vNuevoSujeto)
    strApellido = StringVar(vNuevoSujeto)
    strEdad = StringVar(vNuevoSujeto)
    strEdadDecimal = StringVar(vNuevoSujeto)
    strFechaActual = StringVar(vNuevoSujeto)
    strFechaActual.set(FechaHoy)

    strAreaTrabajo = StringVar(vNuevoSujeto)
    strEmpresa = StringVar(vNuevoSujeto)
    strPeso = IntVar(vNuevoSujeto)
    strEstatura = IntVar(vNuevoSujeto)
    strAfectaciones = StringVar(vNuevoSujeto)
    strObservaciones = StringVar(vNuevoSujeto)

    strNombreEvaluador = StringVar(vNuevoSujeto)
    strApellidosEvaluador = StringVar(vNuevoSujeto)
    strNumeroIdentificacion = IntVar(vNuevoSujeto)
    strNumeroContacto = IntVar(vNuevoSujeto)
    
    iSexo = StringVar(vNuevoSujeto)
    iSexo.set("Masculino")
    iLadoDominante = StringVar(vNuevoSujeto)
    iLadoDominante.set("Derecho")
    iTipoInvestigacion = StringVar(vNuevoSujeto)
    iTipoInvestigacion.set("Clinica")
    


    vNuevoSujeto.wm_title("Nuevo sujeto/paciente")
    vNuevoSujeto.columnconfigure(0,weight=1)
    vNuevoSujeto.rowconfigure(0,weight=1)

    iDia = list(range(1,32))
    iMes = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"]
    iAño = list(range(1900,int(FechaHoy.strftime("%Y"))+1))
    iAño.sort(reverse = True)
    iEvaluadorExistente = []
    iSujetoExistente = []
    
    for x in dbTbEvaluadoresClinico:
        iEvaluadorExistente.append(x['NombreEvaluador'] + " " + x['ApellidosEvaluador'])
    for y in dbTbSujetoClinico:
        iSujetoExistente.append(y['Nombre'] + " " + y['Apellido'])

    global Seudonimo
    global Nombre
    global Apellido
    global Sexo
    global DiaNacimiento
    global MesNacimiento
    global IndexMesNacimiento
    global AñoNacimiento
    global Edad
    global EdadDecimal
    global LadoDominante
    global AreaTrabajo
    global Empresa
    global Peso
    global Estatura
    global Afectaciones
    global Observaciones

    global NombreEvaluador
    global ApellidosEvaluador
    global NumeroIdentificacion
    global NumeroContacto
    global TipoEvaluacion
    global FechaEvaluacion
    global EvaluadorExistente
    
    def clrEntSujeto(event):        
        if (entSujeto.get()=="Seudónimo"):
            entSujeto.delete(0, "end")
        

    def salirNuevoSujeto():
        vNuevoSujeto.destroy()
       
    def getDatos():
        global flagTipoInvestigacion
        if(NewExistSeleccion==1):
            if(flagTipoInvestigacion==True):
                Nombre = strNombre.get()
                Apellido = strApellido.get()
            else:
                Seudonimo = entSujeto.get()
                Nombre = strNombre.get()
                Apellido = strApellido.get()
        else:
            TMPDatosSujeto = OpcionSujetoExistente.get().split(" ")
            Nombre = TMPDatosSujeto[0]
            Apellido = (" ").join(TMPDatosSujeto[1:])
        Sexo = iSexo.get()
        DiaNacimiento = diaOpcion.get()
        MesNacimiento = mesOpcion.get()
        IndexMesNacimiento = iMes.index(MesNacimiento) + 1
        AñoNacimiento = añoOpcion.get()
        FechaNacimiento = date(int(AñoNacimiento),int(IndexMesNacimiento),int(DiaNacimiento))
        Edad = relativedelta(FechaHoy, FechaNacimiento)
        EdadDecimal = round((Edad.years * 12 + Edad.months + Edad.days/30)/12, 3)
        LadoDominante = iLadoDominante.get()
        AreaTrabajo = entAreaTrabajo.get()
        Empresa = entEmpresa.get()
        Peso = entPeso.get()
        Estatura = entEstatura.get()
        Afectaciones = entAfectaciones.get()
        Observaciones = entObservaciones.get()

        NombreEvaluador = entNombreEvaluador.get()
        ApellidosEvaluador = entApellidosEvaluador.get()
        NumeroIdentificacion = entNumeroIdentificacion.get()
        NumeroContacto = entNumeroContacto.get()
        TipoEvaluacion = iTipoInvestigacion.get()
        FechaEvaluacion = entFechaActual.get()
        #EvaluadorExistente = OpcionEvaluadorExistente.get()

        DatosPaciente = {}
        if(flagTipoInvestigacion==True):
            DatosPaciente['Nombre'] = Nombre
            DatosPaciente['Apellido'] = Apellido
        else:
            DatosPaciente['Seudonimo'] = Seudonimo
            DatosPaciente['Nombre'] = Nombre
            DatosPaciente['Apellido'] = Apellido
        DatosPaciente['Sexo'] = Sexo
        DatosPaciente['DiaNacimiento'] = DiaNacimiento
        DatosPaciente['MesNacimiento'] = MesNacimiento
        DatosPaciente['AnoNacimiento'] = AñoNacimiento
        DatosPaciente['Edad'] = str(Edad.years) + " A " +str(Edad.months)+ " M " + str(Edad.days) +" D"
        DatosPaciente['EdadDecimal'] = EdadDecimal 
        DatosPaciente['LadoDominante'] = LadoDominante
        DatosPaciente['AreaTrabajo'] = AreaTrabajo
        DatosPaciente['Empresa'] = Empresa
        DatosPaciente['Peso'] = Peso
        DatosPaciente['Estatura'] = Estatura
        DatosPaciente['Afectaciones'] = Afectaciones
        DatosPaciente['Observaciones'] = Observaciones
        
        DatosEvaluador = {}
        DatosEvaluador['NombreEvaluador'] = NombreEvaluador
        DatosEvaluador['ApellidosEvaluador'] = ApellidosEvaluador
        DatosEvaluador['NumeroIdentificacion'] = NumeroIdentificacion
        DatosEvaluador['NumeroContacto'] = NumeroContacto
        DatosEvaluador['TipoEvaluacion'] = TipoEvaluacion
        DatosEvaluador['FechaEvaluacion'] = FechaEvaluacion

        DatosPacEval = {}
        DatosPacEval['Nombre'] = Nombre
        DatosPacEval['Apellido'] = Apellido
        DatosPacEval['NombreEvaluador'] = NombreEvaluador
        DatosPacEval['ApellidosEvaluador'] = ApellidosEvaluador

        if(flagTipoInvestigacion==True):
            if(NewExistSeleccion==1):
                NameSujeto = Nombre
                LastSujeto = Apellido
            else:
                DatosSujeto = OpcionSujetoExistente.get().split(" ")
                NameSujeto = DatosSujeto[0]
                LastSujeto = (" ").join(DatosSujeto[1:])
            QNombreSujeto = Query()
            dbTbSujetoClinico.upsert(DatosPaciente,((QNombreSujeto.Nombre == NameSujeto)&(QNombreSujeto.Apellido == LastSujeto)))
            QNombreEvaluador = Query()
            dbTbEvaluadoresClinico.upsert(DatosEvaluador,((QNombreEvaluador.NombreEvaluador == NombreEvaluador)&(QNombreEvaluador.ApellidosEvaluador == ApellidosEvaluador)))
            dbTbSujEvalClinico.upsert(DatosPacEval,((QNombreSujeto.Nombre == Nombre)&(QNombreEvaluador.NombreEvaluador == NombreEvaluador)))
        else:
            dbTbSujetoInvestigacion.insert(DatosPaciente)
            dbTbEvaluadoresInvestigacion.insert(DatosEvaluador)
            dbTbSujEvalInvestigacion.insert(DatosPacEval)

        strDatoNombre.set(Nombre)
        strDatoApellido.set(Apellido)
        strDatoSujeto.set(entSujeto.get())
        flagTipoInvestigacion=True

        vNuevoSujeto.destroy()
        
    def CalcEdad():
        DiaNacimiento = diaOpcion.get()
        MesNacimiento = mesOpcion.get()
        iMes = ["Enero","Febrero","Marzo","Abril","Mayo","Junio","Julio","Agosto","Setiembre","Octubre","Noviembre","Diciembre"]
        IndexMesNacimiento = iMes.index(MesNacimiento) + 1
        AñoNacimiento = añoOpcion.get()
        if(DiaNacimiento!="Día" and MesNacimiento!="Mes" and AñoNacimiento!="Año"):
            FechaNacimiento = date(int(AñoNacimiento),int(IndexMesNacimiento),int(DiaNacimiento))
            Edad = relativedelta(FechaHoy, FechaNacimiento)
            EdadDecimal = round((Edad.years * 12 + Edad.months + Edad.days/30)/12, 3)
            strEdadDecimal.set(str(EdadDecimal) + " Años")
            strEdad.set(str(Edad.years) + " años " +str(Edad.months)+ " meses " + str(Edad.days) +" días")
        else:
            print("Error: Completar informacion")
            
    def btnTipoInvestigacion():
        global flagTipoInvestigacion
        global flagTipoExportar
        if(iTipoInvestigacion.get()=="Investigacion"):
            entSujeto.configure(state="normal",foreground="black")
            flagTipoInvestigacion=False
            flagTipoExportar=False
        else:
            entSujeto.delete(0, "end")
            entSujeto.insert(0, 'Seudónimo')
            entSujeto.configure(state="disable",foreground="gray")
            flagTipoInvestigacion=True
            flagTipoExportar=True

                
    def DatosEvaluadorExistente(event):
        DatosEvaluador = OpcionEvaluadorExistente.get().split(" ")
        NameEvaluador = DatosEvaluador[0]
        LastEvaluador = (" ").join(DatosEvaluador[1:])
        QNameEvaluador = Query()
        EvaluadorSeleccionado = dbTbEvaluadoresClinico.search((QNameEvaluador.NombreEvaluador == NameEvaluador)&(QNameEvaluador.ApellidosEvaluador == LastEvaluador))
        strNombreEvaluador.set(EvaluadorSeleccionado[0]["NombreEvaluador"])
        strApellidosEvaluador.set(EvaluadorSeleccionado[0]["ApellidosEvaluador"])
        strNumeroIdentificacion.set(EvaluadorSeleccionado[0]["NumeroIdentificacion"])
        strNumeroContacto.set(EvaluadorSeleccionado[0]["NumeroContacto"])

    def DatosSujetoExistente(event):
        global flagTipoInvestigacion
        flagTipoInvestigacion = True
        DatosSujeto = OpcionSujetoExistente.get().split(" ")
        NameSujeto = DatosSujeto[0]
        LastSujeto = (" ").join(DatosSujeto[1:])
        QNameSujeto = Query()
        SujetoSeleccionado = dbTbSujetoClinico.search((QNameSujeto.Nombre == NameSujeto)&(QNameSujeto.Apellido == LastSujeto))
        strAreaTrabajo.set(SujetoSeleccionado[0]["AreaTrabajo"])
        strEmpresa.set(SujetoSeleccionado[0]["Empresa"])
        strPeso.set(SujetoSeleccionado[0]["Peso"])
        strEstatura.set(SujetoSeleccionado[0]["Estatura"])
        strAfectaciones.set(SujetoSeleccionado[0]["Afectaciones"])
        strObservaciones.set(SujetoSeleccionado[0]["Observaciones"])
        diaOpcion.set(SujetoSeleccionado[0]["DiaNacimiento"])
        mesOpcion.set(SujetoSeleccionado[0]["MesNacimiento"])
        añoOpcion.set(SujetoSeleccionado[0]["AnoNacimiento"])
        iSexo.set(SujetoSeleccionado[0]["Sexo"])
        iLadoDominante.set(SujetoSeleccionado[0]["LadoDominante"])
        
#############################################Datos del Paciente#####################################################################
    lblTipoEvaluacion = ttk.Label(vNuevoSujeto, text="Tipo investigación")
    lblTipoEvaluacion.grid(row=0, column=0)
    if(NewExistSeleccion==1):
        rbtnTipoEvaluacion= ttk.Radiobutton(vNuevoSujeto,text="Investigación",variable=iTipoInvestigacion,value="Investigacion",command=btnTipoInvestigacion)
        rbtnTipoEvaluacion.grid(row=0,column=1)
    else:
        rbtnTipoEvaluacion= ttk.Radiobutton(vNuevoSujeto,text="Investigación",variable=iTipoInvestigacion,value="Investigacion",command=btnTipoInvestigacion,state="disable")
        rbtnTipoEvaluacion.grid(row=0,column=1)
    rbtnTipoEvaluacion = ttk.Radiobutton(vNuevoSujeto,text="Clínica",variable=iTipoInvestigacion,value="Clinica",command=btnTipoInvestigacion)
    rbtnTipoEvaluacion.grid(row=0,column=2)
    entSujeto = ttk.Entry(vNuevoSujeto)
    entSujeto.insert(0, 'Seudónimo')
    entSujeto.bind('<FocusIn>', clrEntSujeto)
    entSujeto.grid(row=0,column=3,sticky='we')
    entSujeto.config(foreground = 'grey',state="disabled")

    lblInfoDatosPaciente = ttk.Label(vNuevoSujeto, text="-------------------------------- Datos del Paciente/Sujeto --------------------------------",font=NORM_FONT)
    lblInfoDatosPaciente.grid(row=1, column=0,columnspan=4)
    if(NewExistSeleccion==1):
        lblNombre = ttk.Label(vNuevoSujeto, text="Nombre")
        lblNombre.grid(row=2, column=0)
        entNombre = ttk.Entry(vNuevoSujeto,textvariable=strNombre)
        entNombre.grid(row=2, column=1,sticky='we')
        lblApellidos = ttk.Label(vNuevoSujeto, text="Apellidos")
        lblApellidos.grid(row=2, column=2)
        entApellidos = ttk.Entry(vNuevoSujeto,textvariable=strApellido)
        entApellidos.grid(row=2, column=3,sticky='we')
    else:
        lblSujetoExistente = ttk.Label(vNuevoSujeto, text="Nombre y apellidos")
        lblSujetoExistente.grid(row=2, column=0)
        OpcionSujetoExistente = ttk.Combobox(vNuevoSujeto,values=iSujetoExistente)
        OpcionSujetoExistente.set("")
        OpcionSujetoExistente.grid(row=2, column=1,sticky='we')
        OpcionSujetoExistente.bind("<<ComboboxSelected>>", DatosSujetoExistente)
        
    lblSexo = ttk.Label(vNuevoSujeto, text="Sexo")
    lblSexo.grid(row=3, column=0)
    rbtnSexoM = ttk.Radiobutton(vNuevoSujeto,text="Masculino",variable=iSexo,value="Masculino")
    rbtnSexoM.grid(row=3,column=1)
    rbtnSexoF = ttk.Radiobutton(vNuevoSujeto,text="Femenino",variable=iSexo,value="Femenino")
    rbtnSexoF.grid(row=3,column=2)
    lblFechaNacimiento = ttk.Label(vNuevoSujeto, text="Fecha nacimiento")
    lblFechaNacimiento.grid(row=4, column=0)
    diaOpcion = ttk.Combobox(vNuevoSujeto,values=iDia)
    diaOpcion.set("Día")
    diaOpcion.grid(row=4, column=1)
    mesOpcion = ttk.Combobox(vNuevoSujeto,values=iMes)
    mesOpcion.set("Mes")
    mesOpcion.grid(row=4, column=2)
    añoOpcion = ttk.Combobox(vNuevoSujeto,values=iAño)
    añoOpcion.set("Año")
    añoOpcion.grid(row=4, column=3)
    lblInfoEdad = ttk.Label(vNuevoSujeto, text="Edad exacta")
    lblInfoEdad.grid(row=5, column=0)
    entEdad = ttk.Entry(vNuevoSujeto,textvariable=strEdad,state="readonly")
    entEdad.grid(row=5, column=1,sticky='we')
    entEdadDecimal = ttk.Entry(vNuevoSujeto,textvariable=strEdadDecimal,state="readonly")
    entEdadDecimal.grid(row=5, column=2,sticky='we')
    btnCalcEdad = ttk.Button(vNuevoSujeto, text = "Calcular edad", command = CalcEdad)
    btnCalcEdad.grid(row=5,column=3,sticky='we')
    lblLadoDominante = ttk.Label(vNuevoSujeto, text="Lado Dominante")
    lblLadoDominante.grid(row=6, column=0)
    rbtnDerecho = ttk.Radiobutton(vNuevoSujeto,text="Derecho",variable=iLadoDominante,value="Derecho")
    rbtnDerecho.grid(row=6,column=1)
    rbtnIzquierdo = ttk.Radiobutton(vNuevoSujeto,text="Izquierdo",variable=iLadoDominante,value="Izquierdo")
    rbtnIzquierdo.grid(row=6,column=2)
    lblAreaTrabajo = ttk.Label(vNuevoSujeto, text="Area de trabajo")
    lblAreaTrabajo.grid(row=7,column=0)
    entAreaTrabajo = ttk.Entry(vNuevoSujeto,textvariable=strAreaTrabajo)
    entAreaTrabajo.grid(row=7, column=1,sticky='we')
    lblEmpresa = ttk.Label(vNuevoSujeto, text="Empresa")
    lblEmpresa.grid(row=7,column=2)
    entEmpresa = ttk.Entry(vNuevoSujeto,textvariable=strEmpresa)
    entEmpresa.grid(row=7, column=3,sticky='we')
    lblPeso = ttk.Label(vNuevoSujeto, text="Peso")
    lblPeso.grid(row=8,column=0)
    entPeso = ttk.Entry(vNuevoSujeto,textvariable=strPeso)
    entPeso.grid(row=8, column=1,sticky='we')
    lblEstatura = ttk.Label(vNuevoSujeto, text="Estatura")
    lblEstatura.grid(row=8,column=2)
    entEstatura = ttk.Entry(vNuevoSujeto,textvariable=strEstatura)
    entEstatura.grid(row=8, column=3,sticky='we')
    lblAfectaciones = ttk.Label(vNuevoSujeto, text="Afectaciones")
    lblAfectaciones.grid(row=9,column=0)
    entAfectaciones = ttk.Entry(vNuevoSujeto,textvariable=strAfectaciones)
    entAfectaciones.grid(row=9,column=1,sticky='we',columnspan=3)
    lblObservaciones = ttk.Label(vNuevoSujeto, text="Observaciones")
    lblObservaciones.grid(row=10,column=0)
    entObservaciones = ttk.Entry(vNuevoSujeto,textvariable=strObservaciones)
    entObservaciones.grid(row=10,column=1,sticky='we',columnspan=3)
#############################################Datos del evaluador#####################################################################
    lblInfoDatosEvaluador = ttk.Label(vNuevoSujeto, text="-------------------------------- Datos del evaluador --------------------------------")
    lblInfoDatosEvaluador.grid(row=11, column=0,columnspan=4)
    lblNombreEvaluador = ttk.Label(vNuevoSujeto, text="Nombre")
    lblNombreEvaluador.grid(row=12, column=0)
    entNombreEvaluador = ttk.Entry(vNuevoSujeto,textvariable=strNombreEvaluador)
    entNombreEvaluador.grid(row=12, column=1,sticky='we')
    lblApellidosEvaluador = ttk.Label(vNuevoSujeto, text="Apellidos")
    lblApellidosEvaluador.grid(row=12, column=2)
    entApellidosEvaluador = ttk.Entry(vNuevoSujeto,textvariable=strApellidosEvaluador)
    entApellidosEvaluador.grid(row=12, column=3,sticky='we')
    lblNumeroIdentificacion = ttk.Label(vNuevoSujeto, text="# de identificación")
    lblNumeroIdentificacion.grid(row=13, column=0)
    entNumeroIdentificacion = ttk.Entry(vNuevoSujeto,textvariable=strNumeroIdentificacion)
    entNumeroIdentificacion.grid(row=13, column=1,sticky='we')
    lblNumeroContacto = ttk.Label(vNuevoSujeto, text="# de teléfono")
    lblNumeroContacto.grid(row=13, column=2)
    entNumeroContacto = ttk.Entry(vNuevoSujeto,textvariable=strNumeroContacto)
    entNumeroContacto.grid(row=13, column=3,sticky='we')
    lblFechaActual = ttk.Label(vNuevoSujeto, text="Fecha")
    lblFechaActual.grid(row=14, column=0)
    entFechaActual = ttk.Entry(vNuevoSujeto,textvariable=strFechaActual)
    entFechaActual.grid(row=14, column=1,sticky='we')
    lblEvaluadorExistente = ttk.Label(vNuevoSujeto, text="Evaluador existente")
    lblEvaluadorExistente.grid(row=14, column=2)
    OpcionEvaluadorExistente = ttk.Combobox(vNuevoSujeto,values=iEvaluadorExistente)
    OpcionEvaluadorExistente.set("Evaluador")
    OpcionEvaluadorExistente.grid(row=14, column=3)
    OpcionEvaluadorExistente.bind("<<ComboboxSelected>>", DatosEvaluadorExistente)


    btnAceptar = ttk.Button(vNuevoSujeto, text = "Aceptar", command = getDatos)
    btnAceptar.grid(row=15,sticky='we',column=1)
    btnSalir = ttk.Button(vNuevoSujeto, text = "Cancelar", command = salirNuevoSujeto)
    btnSalir.grid(row=15,sticky='we',column=2)
    
    vNuevoSujeto.resizable(width=False, height=False)
    vNuevoSujeto.mainloop()
 

#########################################################################
#                                                                       #
#               Clases para crear las ventanas del programa             #
#                                                                       #
#########################################################################

class Principal(tk.Tk):

    def __init__(self, *args, **kwargs):
        
        tk.Tk.__init__(self, *args, **kwargs)
        tk.Tk.wm_title(self, "Biomec ErgoTec")
        
        pConteiner = tk.Frame(self)
        pConteiner.pack(side="top", fill="both", expand = True)
        pConteiner.grid_rowconfigure(0, weight=1)
        pConteiner.grid_columnconfigure(0, weight=1)

        BarraMenu = tk.Menu(pConteiner)
        
        mGrabacion = tk.Menu(BarraMenu, tearoff=0)
        mGrabacion.add_command(label="Nuevo", command = lambda: NuevoSujeto(1))
        mGrabacion.add_separator()
        mGrabacion.add_command(label="Existente", command = lambda: NuevoSujeto(2))
        mGrabacion.add_separator()
        mGrabacion.add_command(label="Salir", command=quit)
        BarraMenu.add_cascade(label="Grabación", menu=mGrabacion)

        mCalibracion = tk.Menu(BarraMenu, tearoff=0)
        mCalibracion.add_command(label="Plataforma de fuerza", command = lambda: Calibracion())
        mCalibracion.add_separator()
        mCalibracion.add_command(label="Electrogniómetro", command = lambda: None)
        BarraMenu.add_cascade(label="Calibración", menu=mCalibracion)

        mInformacion = tk.Menu(BarraMenu, tearoff=0)
        mInformacion.add_command(label="Plataforma de fuerza", command = lambda: InformacionPlataformaFuerza())
        mInformacion.add_separator()
        mInformacion.add_command(label="Electrogniómetro", command = lambda: None)
        mInformacion.add_separator()
        mInformacion.add_command(label="ErgoTec", command = lambda: InformacionErgoTec())
        BarraMenu.add_cascade(label="Acerca de", menu=mInformacion)

        mAyuda = tk.Menu(BarraMenu, tearoff=0)
        mAyuda.add_command(label="Plataforma de fuerza", command = lambda: popupmsg("Not supported just yet!"))
        mAyuda.add_separator()
        mAyuda.add_command(label="Electrogniómetro", command = lambda: None)
        BarraMenu.add_cascade(label="Ayuda", menu=mAyuda)

        tk.Tk.config(self, menu=BarraMenu)

        self.frames = {}
        pVentana = Pagina_Inicio(pConteiner, self)
        self.frames[Pagina_Inicio] = pVentana
        pVentana.grid(row=0, column=0, sticky="nsew")
        self.mostrar_Ventana(Pagina_Inicio)

    def mostrar_Ventana(self, cont):

        pVentana = self.frames[cont]
        pVentana.tkraise()

        
class Pagina_Inicio(tk.Frame):

    def __init__(self, parent, controller):
        tk.Frame.__init__(self,parent)
        frame_P1 = tk.Frame(self)
        frame_P1.pack(fill="x")
        frame_P2 = tk.Frame(frame_P1)
        frame_P2.pack(fill="x",side=tk.TOP)
        frame_P3 = tk.Frame(frame_P1)
        frame_P3.pack(side=tk.RIGHT,fill=tk.Y)
        
        lblTitulo = tk.Label(frame_P2, text="Plataforma de fuerza", font=LARGE_FONT)
        lblTitulo.pack(side=tk.LEFT,pady=10,padx=10)

        lblSensor1 = tk.Label(frame_P3, text="Sensor 1",padx=10,pady=10, borderwidth=2, relief="groove", font=NORM_FONT)
        lblSensor1.grid(row=0, column=0,sticky='we')
        lblSensor2 = tk.Label(frame_P3, text="Sensor 2",padx=10,pady=10, borderwidth=2, relief="groove", font=NORM_FONT)
        lblSensor2.grid(row=0, column=1,sticky='we')
        lblSensor3 = tk.Label(frame_P3, text="Sensor 3",padx=10,pady=10, borderwidth=2, relief="groove", font=NORM_FONT)
        lblSensor3.grid(row=0, column=2,sticky='we')
        lblSensor4 = tk.Label(frame_P3, text="Sensor 4",padx=10,pady=10, borderwidth=2, relief="groove", font=NORM_FONT)
        lblSensor4.grid(row=0, column=3,sticky='we')

        global strDatoSensor1
        strDatoSensor1 = StringVar(frame_P3)
        global strDatoSensor2
        strDatoSensor2 = StringVar(frame_P3)
        global strDatoSensor3
        strDatoSensor3 = StringVar(frame_P3)
        global strDatoSensor4
        strDatoSensor4 = StringVar(frame_P3)

        entSensor1 = ttk.Entry(frame_P3,textvariable=strDatoSensor1,state="readonly",justify='center')
        entSensor1.grid(row=1, column=0,sticky='we')
        entSensor2 = ttk.Entry(frame_P3,textvariable=strDatoSensor2,state="readonly",justify='center')
        entSensor2.grid(row=1, column=1,sticky='we')
        entSensor3 = ttk.Entry(frame_P3,textvariable=strDatoSensor3,state="readonly",justify='center')
        entSensor3.grid(row=1, column=2,sticky='we')
        entSensor4 = ttk.Entry(frame_P3,textvariable=strDatoSensor4,state="readonly",justify='center')
        entSensor4.grid(row=1, column=3,sticky='we')

        global strDatoFuerza
        strDatoFuerza = StringVar(frame_P3)
        lblFuerza = tk.Label(frame_P3, text="Fuerza: ",padx=10,pady=11, borderwidth=2, relief="groove", font=NORM_FONT)
        lblFuerza.grid(row=2, column=0,sticky='we')
        entFuerza = ttk.Entry(frame_P3,textvariable=strDatoFuerza,state="readonly",justify='center')
        entFuerza.grid(row=2, column=1,sticky='we', ipadx=10, ipady=10)

        global strDatoFiltrado
        strDatoFiltrado = StringVar(frame_P3)
        lblFiltro = tk.Label(frame_P3, text="Dato filtrado: ",padx=10,pady=11, borderwidth=2, relief="groove", font=NORM_FONT)
        lblFiltro.grid(row=2, column=2,sticky='we')
        entFiltro = ttk.Entry(frame_P3,textvariable=strDatoFiltrado,state="readonly",justify='center')
        entFiltro.grid(row=2, column=3,sticky='we', ipadx=10, ipady=10)


        btnStopPrev = ttk.Button(frame_P3, text = "Stop Visualización", command = StopGrafico)
        btnStopPrev.grid(row=3,sticky='we',column=0, ipady=2)
        btnPlayPrev = ttk.Button(frame_P3, text = "PlayVisualización", command = PlayGrafico)
        btnPlayPrev.grid(row=3,sticky='we',column=1, ipady=2)

        global strBtnGrabacion
        strBtnGrabacion = StringVar(frame_P3)
        strBtnGrabacion.set("Iniciar Grabación")
        btnStop = ttk.Button(frame_P3,textvariable = strBtnGrabacion, command = FlagGrabacion)
        btnStop.grid(row=3,sticky='we',column=2,columnspan=2, ipady=2)

        strDatoGrafico = StringVar(frame_P3)
        lblInfoGrafico = tk.Label(frame_P3, text="Comentario: ",padx=10,pady=10, borderwidth=2, relief="groove", font=NORM_FONT)
        lblInfoGrafico.grid(row=4, column=0,sticky='we')
        entInfoGrafico = ttk.Entry(frame_P3,textvariable=strDatoGrafico,font=NORM_FONT)
        entInfoGrafico.grid(row=4, column=1,sticky='we',columnspan=3, ipadx=10, ipady=8)

        global strDatoNombre
        strDatoNombre = StringVar(frame_P3)
        lblInfoNombre = tk.Label(frame_P3, text="Nombre: ",padx=5,pady=6, borderwidth=2, relief="groove", font=NORM_FONT)
        lblInfoNombre.grid(row=5, column=0,sticky='we')
        entInfoNombre = ttk.Entry(frame_P3,textvariable=strDatoNombre,font=NORM_FONT,state="readonly")
        entInfoNombre.grid(row=5, column=1,sticky='we', ipadx=5, ipady=5)

        global strDatoApellido
        strDatoApellido = StringVar(frame_P3)
        lblInfoApellido = tk.Label(frame_P3, text="Apellido: ",padx=5,pady=6, borderwidth=2, relief="groove", font=NORM_FONT)
        lblInfoApellido.grid(row=5, column=2,sticky='we')
        entInfoApellido = ttk.Entry(frame_P3,textvariable=strDatoApellido,font=NORM_FONT,state="readonly")
        entInfoApellido.grid(row=5, column=3,sticky='we', ipadx=5, ipady=5)

        btnMostarGrafico = ttk.Button(frame_P3,text = "Mostrar Grafico", command = MostrarGrafico)
        btnMostarGrafico.grid(row=6,sticky='we',column=0,columnspan=2, ipady=2)
        btnExportarDatos = ttk.Button(frame_P3,text = "Exportar Datos", command = ExportarDatos)
        btnExportarDatos.grid(row=6,sticky='we',column=2,columnspan=2, ipady=2)

        btnSetCero = ttk.Button(frame_P3,text = "Establecer cero", command = EstablecerCero)
        btnSetCero.grid(row=7,sticky='we',column=1,columnspan=2, ipady=2)
        
        
        canvas = FigureCanvasTkAgg(fig, frame_P1)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.X, expand=True)
        
        toolbar = NavigationToolbar2Tk(canvas, frame_P1)
        toolbar.update()
        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.X, expand=True)

        mplcursors.cursor(multiple=True).connect("add", lambda sel: sel.annotation.set_text(strDatoGrafico.get()))
###########################################Electrogoniometro################################################
        frame_E1 = tk.Frame(self)
        frame_E1.pack(fill="x")
        frame_E2 = tk.Frame(frame_E1)
        frame_E2.pack(fill="x",side=tk.TOP)
        frame_E3 = tk.Frame(frame_E1, width=545)
        frame_E3.pack(side=tk.RIGHT,fill=tk.Y)
        lblTitulo2 = tk.Label(frame_E2, text="Electrogoniómetro", font=LARGE_FONT)
        lblTitulo2.pack(side=tk.LEFT,pady=10,padx=10)
        
        canvas = FigureCanvasTkAgg(fig2, frame_E1)
        canvas.draw()
        canvas.get_tk_widget().pack(side=tk.TOP, fill=tk.X, expand=True)
        
        toolbar = NavigationToolbar2Tk(canvas, frame_E1)
        toolbar.update()
        canvas._tkcanvas.pack(side=tk.TOP, fill=tk.X, expand=True)

        if(Embebido==False):
            messagebox.showwarning("Warning","PSOC no conectado\n\nSe iniciará protocolo de prueba")


app = Principal()
app.geometry("1280x720")
ani = animation.FuncAnimation(fig,AnimarGrafico,fargs=(ys,),blit=True)
app.mainloop()
